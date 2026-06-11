#!/usr/bin/env python
"""
06_merge_summary.py
===================
Reads the bird_orthologues_domains.xlsx and produces a clean color-coded
summary table with only the columns you need. No API calls.

Usage:
  python 06_merge_summary.py results/human_cell_death_gene_universe_v1_bird_orthologues_domains.xlsx
"""

import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path


def status_call(pct_id, otype):
    if pct_id is None:
        return "ABSENT"
    try:
        pct = float(pct_id)
    except (ValueError, TypeError):
        return "ABSENT"
    if pct < 10:
        return "NOISE (<10%)"
    if pct < 20:
        return "DUBIOUS (10-20%)"
    return "PRESENT"


def status_fill(status):
    green = PatternFill("solid", fgColor="C6EFCE")
    red = PatternFill("solid", fgColor="FFC7CE")
    yellow = PatternFill("solid", fgColor="FFEB9C")
    if "PRESENT" in status:
        return green
    if "NOISE" in status or "DUBIOUS" in status:
        return yellow
    return red


def verdict_fill(verdict):
    green = PatternFill("solid", fgColor="C6EFCE")
    light_green = PatternFill("solid", fgColor="E2EFDA")
    yellow = PatternFill("solid", fgColor="FFEB9C")
    orange = PatternFill("solid", fgColor="FCE4D6")
    red = PatternFill("solid", fgColor="FFC7CE")
    grey = PatternFill("solid", fgColor="D9D9D9")
    if verdict == "FULL_MATCH":
        return green
    if verdict == "GOOD_MATCH":
        return light_green
    if verdict == "PARTIAL_MATCH":
        return yellow
    if verdict == "WEAK_MATCH":
        return orange
    if verdict == "NO_MATCH":
        return red
    return grey  # NO_ORTHOLOGUE, NO_UNIPROT


def main():
    if len(sys.argv) < 2:
        print("Usage: python 06_merge_summary.py <domains.xlsx>")
        sys.exit(1)

    input_path = Path(sys.argv[1])
    output_path = input_path.parent / "orthologue_domain_summary.xlsx"

    wb_in = openpyxl.load_workbook(input_path)
    ws_in = wb_in.active

    # Input column indices (1-based, confirmed)
    IN = {
        "symbol": 1,
        "chk_pct": 14,
        "chk_type": 11,
        "duck_pct": 21,
        "duck_type": 18,
        "chk_verdict": 26,
        "chk_match_pct": 27,
        "chk_shared": 28,
        "chk_missing": 29,
        "chk_extra": 30,
        "duck_verdict": 32,
        "duck_match_pct": 33,
        "duck_shared": 34,
        "duck_missing": 35,
        "duck_extra": 36,
    }

    # Build output
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Summary"

    out_headers = [
        "Gene Symbol",
        "Chicken % Identity",
        "Chicken Orthologue Type",
        "Chicken Status",
        "Duck % Identity",
        "Duck Orthologue Type",
        "Duck Status",
        "Chicken Domain Verdict",
        "Chicken Domain Match %",
        "Chicken Shared Domains",
        "Chicken Missing Domains",
        "Chicken Extra Domains",
        "Duck Domain Verdict",
        "Duck Domain Match %",
        "Duck Shared Domains",
        "Duck Missing Domains",
        "Duck Extra Domains",
    ]

    hfont = Font(bold=True, size=10, name="Arial")
    hfill = PatternFill("solid", fgColor="D9E1F2")
    wrap = Alignment(wrap_text=True, vertical="top")
    body = Font(size=9, name="Arial")

    for col, h in enumerate(out_headers, 1):
        cell = ws_out.cell(row=1, column=col, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = wrap

    total = ws_in.max_row - 1

    for row in range(2, ws_in.max_row + 1):
        def val(key):
            return ws_in.cell(row=row, column=IN[key]).value

        symbol = val("symbol") or ""
        chk_pct = val("chk_pct")
        chk_type = val("chk_type") or ""
        duck_pct = val("duck_pct")
        duck_type = val("duck_type") or ""
        chk_verdict = val("chk_verdict") or "NO_ORTHOLOGUE"
        chk_match_pct = val("chk_match_pct")
        chk_shared = val("chk_shared") or ""
        chk_missing = val("chk_missing") or ""
        chk_extra = val("chk_extra") or ""
        duck_verdict = val("duck_verdict") or "NO_ORTHOLOGUE"
        duck_match_pct = val("duck_match_pct")
        duck_shared = val("duck_shared") or ""
        duck_missing = val("duck_missing") or ""
        duck_extra = val("duck_extra") or ""

        out_row = row
        r = out_row

        # Gene symbol
        ws_out.cell(row=r, column=1, value=symbol).font = body

        # Chicken basic
        ws_out.cell(row=r, column=2, value=chk_pct).font = body
        ws_out.cell(row=r, column=3, value=chk_type).font = body
        chk_status = status_call(chk_pct, chk_type)
        c = ws_out.cell(row=r, column=4, value=chk_status)
        c.font = body
        c.fill = status_fill(chk_status)

        # Duck basic
        ws_out.cell(row=r, column=5, value=duck_pct).font = body
        ws_out.cell(row=r, column=6, value=duck_type).font = body
        duck_status = status_call(duck_pct, duck_type)
        c = ws_out.cell(row=r, column=7, value=duck_status)
        c.font = body
        c.fill = status_fill(duck_status)

        # Chicken domains
        c = ws_out.cell(row=r, column=8, value=chk_verdict)
        c.font = body
        c.fill = verdict_fill(chk_verdict)
        ws_out.cell(row=r, column=9, value=chk_match_pct).font = body
        ws_out.cell(row=r, column=10, value=chk_shared).font = body
        ws_out.cell(row=r, column=11, value=chk_missing).font = body
        ws_out.cell(row=r, column=12, value=chk_extra).font = body

        # Duck domains
        c = ws_out.cell(row=r, column=13, value=duck_verdict)
        c.font = body
        c.fill = verdict_fill(duck_verdict)
        ws_out.cell(row=r, column=14, value=duck_match_pct).font = body
        ws_out.cell(row=r, column=15, value=duck_shared).font = body
        ws_out.cell(row=r, column=16, value=duck_missing).font = body
        ws_out.cell(row=r, column=17, value=duck_extra).font = body

        # Wrap all
        for col in range(1, 18):
            ws_out.cell(row=r, column=col).alignment = wrap

    # Column widths
    widths = [12, 10, 18, 14, 10, 18, 14, 18, 10, 45, 45, 45, 18, 10, 45, 45, 45]
    for i, w in enumerate(widths, 1):
        col_letter = openpyxl.utils.get_column_letter(i)
        ws_out.column_dimensions[col_letter].width = w

    ws_out.freeze_panes = "A2"
    ws_out.auto_filter.ref = f"A1:Q{total + 1}"

    wb_out.save(output_path)
    print(f"Done. {total} genes → {output_path}")


if __name__ == "__main__":
    main()
