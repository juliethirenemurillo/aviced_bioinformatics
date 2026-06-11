#!/usr/bin/env python3
"""
01_verify_uniprot.py
====================
Verify UniProt accessions in the human cell death gene universe table.

For each row, queries the UniProt REST API and checks:
  1. Does the accession exist?
  2. Is it reviewed (Swiss-Prot) or unreviewed (TrEMBL)?
  3. Does the gene name in UniProt match our table?
  4. Is the organism Homo sapiens?
  5. Retrieves: protein name, gene name, organism, length, InterPro domains

Outputs:
  - Updated xlsx with new columns: UP_Status, UP_GeneName, UP_ProteinName,
    UP_Organism, UP_Length, UP_Domains, UP_Match
  - Console report of mismatches

Requirements:
  pip install requests openpyxl

Usage:
  python 01_verify_uniprot.py human_cell_death_gene_universe_v1.xlsx

Run on your laptop — UniProt API is free and needs no authentication.
"""

import sys
import time
import requests
import openpyxl
from pathlib import Path


UNIPROT_API = "https://rest.uniprot.org/uniprotkb"


def fetch_uniprot_entry(accession: str, max_retries: int = 3) -> dict | None:
    """Fetch a single UniProt entry by accession."""
    url = f"{UNIPROT_API}/{accession}.json"
    for attempt in range(max_retries):
        try:
            r = requests.get(url, timeout=15)
            if r.status_code == 200:
                return r.json()
            elif r.status_code == 404:
                return None
            elif r.status_code == 429:
                wait = int(r.headers.get("Retry-After", 5))
                print(f"  Rate limited, waiting {wait}s...")
                time.sleep(wait)
            else:
                print(f"  Unexpected status {r.status_code} for {accession}")
                time.sleep(2)
        except requests.exceptions.RequestException as e:
            print(f"  Request error for {accession}: {e}")
            time.sleep(2)
    return None


def parse_entry(data: dict) -> dict:
    """Extract relevant fields from a UniProt JSON entry."""
    # Review status
    reviewed = "reviewed" if data.get("entryType") == "UniProtKB reviewed (Swiss-Prot)" else "unreviewed"

    # Gene name(s)
    genes = data.get("genes", [])
    gene_name = ""
    if genes:
        primary = genes[0].get("geneName", {})
        gene_name = primary.get("value", "")

    # Protein name
    protein_desc = data.get("proteinDescription", {})
    rec_name = protein_desc.get("recommendedName", {})
    prot_name = rec_name.get("fullName", {}).get("value", "")
    if not prot_name:
        sub_names = protein_desc.get("submissionNames", [])
        if sub_names:
            prot_name = sub_names[0].get("fullName", {}).get("value", "")

    # Organism
    organism = data.get("organism", {}).get("scientificName", "")

    # Sequence length
    seq_length = data.get("sequence", {}).get("length", "")

    # InterPro cross-references
    xrefs = data.get("uniProtKBCrossReferences", [])
    interpro = []
    for xref in xrefs:
        if xref.get("database") == "InterPro":
            props = xref.get("properties", [])
            entry_name = ""
            for p in props:
                if p.get("key") == "EntryName":
                    entry_name = p.get("value", "")
            interpro.append(f"{xref.get('id', '')}:{entry_name}")

    return {
        "status": reviewed,
        "gene_name": gene_name,
        "protein_name": prot_name,
        "organism": organism,
        "length": seq_length,
        "domains": "; ".join(interpro[:15]),  # cap at 15 to avoid cell overflow
    }


def main():
    if len(sys.argv) < 2:
        print("Usage: python 01_verify_uniprot.py <input.xlsx>")
        sys.exit(1)

    input_path = Path(sys.argv[1])
    output_path = input_path.with_stem(input_path.stem + "_verified")

    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    # Find column indices (1-based)
    headers = {cell.value: cell.column for cell in ws[1]}
    col_symbol = headers.get("Gene Symbol")
    col_accession = headers.get("UniProt Accession")

    if not col_symbol or not col_accession:
        print("ERROR: Could not find 'Gene Symbol' or 'UniProt Accession' columns")
        sys.exit(1)

    # Add new columns
    new_cols = ["UP_Status", "UP_GeneName", "UP_ProteinName", "UP_Organism",
                "UP_Length", "UP_Domains", "UP_Match"]
    start_col = ws.max_column + 1
    for i, name in enumerate(new_cols):
        ws.cell(row=1, column=start_col + i, value=name)
        ws.cell(row=1, column=start_col + i).font = openpyxl.styles.Font(bold=True)

    # Process each gene
    mismatches = []
    not_found = []
    total = ws.max_row - 1

    for row in range(2, ws.max_row + 1):
        symbol = ws.cell(row=row, column=col_symbol).value or ""
        accession = ws.cell(row=row, column=col_accession).value or ""
        accession = accession.strip()

        print(f"[{row-1}/{total}] {symbol} ({accession})...", end=" ")

        if not accession:
            print("SKIP (no accession)")
            ws.cell(row=row, column=start_col + 6, value="NO_ACCESSION")
            continue

        data = fetch_uniprot_entry(accession)

        if data is None:
            print("NOT FOUND")
            ws.cell(row=row, column=start_col + 6, value="NOT_FOUND")
            not_found.append((symbol, accession))
            continue

        info = parse_entry(data)

        ws.cell(row=row, column=start_col + 0, value=info["status"])
        ws.cell(row=row, column=start_col + 1, value=info["gene_name"])
        ws.cell(row=row, column=start_col + 2, value=info["protein_name"])
        ws.cell(row=row, column=start_col + 3, value=info["organism"])
        ws.cell(row=row, column=start_col + 4, value=info["length"])
        ws.cell(row=row, column=start_col + 5, value=info["domains"])

        # Check match
        issues = []
        if info["organism"] != "Homo sapiens":
            issues.append(f"ORGANISM:{info['organism']}")
        if info["gene_name"].upper() != symbol.upper():
            # Check aliases too
            issues.append(f"GENE:{info['gene_name']}")
        if info["status"] != "reviewed":
            issues.append("UNREVIEWED")

        match_status = "OK" if not issues else "|".join(issues)
        ws.cell(row=row, column=start_col + 6, value=match_status)

        if issues:
            mismatches.append((symbol, accession, match_status))
            print(f"ISSUES: {match_status}")
        else:
            print("OK")

        # Be polite to the API
        time.sleep(0.5)

    # Save
    wb.save(output_path)
    print(f"\n{'='*60}")
    print(f"Saved: {output_path}")
    print(f"Total genes: {total}")
    print(f"Not found: {len(not_found)}")
    print(f"Mismatches: {len(mismatches)}")

    if not_found:
        print(f"\n--- NOT FOUND ---")
        for sym, acc in not_found:
            print(f"  {sym}: {acc}")

    if mismatches:
        print(f"\n--- MISMATCHES ---")
        for sym, acc, issues in mismatches:
            print(f"  {sym} ({acc}): {issues}")


if __name__ == "__main__":
    main()
