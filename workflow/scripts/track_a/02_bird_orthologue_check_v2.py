#!/usr/bin/env python3
"""
02_bird_orthologue_check.py (FIXED)
====================================
For each human gene in the table, query Ensembl Compara for orthologues in:
  - Chicken (Gallus gallus)
  - Duck (Anas platyrhynchos) change for anas_platyrhynchos_platyrhynchos

Requirements:
  pip install requests openpyxl

Usage:
  python 02_bird_orthologue_check.py human_cell_death_gene_universe_v1.xlsx
"""

import sys
import time
import requests
import openpyxl
from pathlib import Path


ENSEMBL_REST = "https://rest.ensembl.org"

TARGET_SPECIES = [
    {"name": "chicken", "ensembl": "gallus_gallus"},
    {"name": "duck", "ensembl": "anas_platyrhynchos_platyrhynchos"},
]


def ensembl_get(endpoint: str, params: dict = None, max_retries: int = 3):
    url = f"{ENSEMBL_REST}{endpoint}"
    headers = {"Accept": "application/json"}
    for attempt in range(max_retries):
        try:
            r = requests.get(url, headers=headers, params=params, timeout=20)
            if r.status_code == 200:
                return r.json()
            elif r.status_code == 429:
                wait = float(r.headers.get("Retry-After", 2))
                print(f"    Rate limited, waiting {wait}s...")
                time.sleep(wait)
            elif r.status_code == 404:
                return None
            else:
                print(f"    Unexpected status {r.status_code}")
                time.sleep(1)
        except requests.exceptions.RequestException as e:
            print(f"    Request error: {e}")
            time.sleep(2)
    return None


def lookup_human_gene(symbol: str) -> str | None:
    data = ensembl_get(f"/lookup/symbol/homo_sapiens/{symbol}")
    if data and "id" in data:
        return data["id"]
    return None


def get_orthologues(ensembl_id: str, target_species: str) -> list[dict]:
    # FIX: include homo_sapiens in the path
    data = ensembl_get(
        f"/homology/id/homo_sapiens/{ensembl_id}",
        params={"target_species": target_species, "type": "orthologues"}
    )
    if not data:
        return []

    homologies = data.get("data", [{}])[0].get("homologies", [])
    results = []
    for h in homologies:
        target = h.get("target", {})
        source = h.get("source", {})
        results.append({
            "type": h.get("type", "unknown"),
            "target_id": target.get("id", ""),
            "target_symbol": target.get("gene_member_id", ""),
            "target_protein": target.get("protein_id", ""),
            "pct_id_query": source.get("perc_id", ""),
            "pct_id_target": target.get("perc_id", ""),
            "pct_pos": target.get("perc_pos", ""),
            "dn_ds": h.get("dn_ds", ""),
            "is_high_confidence": h.get("is_high_confidence", ""),
            "taxonomy_level": h.get("taxonomy_level", ""),
        })
    return results


def try_aliases(symbol: str, aliases: str) -> str | None:
    eid = lookup_human_gene(symbol)
    if eid:
        return eid
    if aliases:
        for alias in aliases.split(","):
            alias = alias.strip()
            if alias:
                eid = lookup_human_gene(alias)
                if eid:
                    print(f"    (found via alias: {alias})")
                    return eid
    return None


def main():
    if len(sys.argv) < 2:
        print("Usage: python 02_bird_orthologue_check.py <input.xlsx>")
        sys.exit(1)

    input_path = Path(sys.argv[1])
    output_path = input_path.with_stem(input_path.stem + "_bird_orthologues")

    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    headers = {cell.value: cell.column for cell in ws[1]}
    col_symbol = headers.get("Gene Symbol")
    col_aliases = headers.get("Aliases")

    if not col_symbol:
        print("ERROR: Could not find 'Gene Symbol' column")
        sys.exit(1)

    # Add new columns for each target species
    start_col = ws.max_column + 1
    col_map = {}
    for sp in TARGET_SPECIES:
        name = sp["name"]
        cols_for_sp = [
            f"{name}_orthologue_type",
            f"{name}_target_id",
            f"{name}_target_symbol",
            f"{name}_pct_id",
            f"{name}_pct_pos",
            f"{name}_confidence",
            f"{name}_dn_ds",
        ]
        col_map[name] = {}
        for cname in cols_for_sp:
            ws.cell(row=1, column=start_col, value=cname)
            ws.cell(row=1, column=start_col).font = openpyxl.styles.Font(bold=True)
            col_map[name][cname] = start_col
            start_col += 1

    total = ws.max_row - 1
    summary = {sp["name"]: {"present": 0, "absent": 0, "not_in_ensembl": 0} for sp in TARGET_SPECIES}

    for row in range(2, ws.max_row + 1):
        symbol = (ws.cell(row=row, column=col_symbol).value or "").strip()
        aliases = (ws.cell(row=row, column=col_aliases).value or "").strip() if col_aliases else ""

        print(f"\n[{row-1}/{total}] {symbol}")

        ensembl_id = try_aliases(symbol, aliases)

        if not ensembl_id:
            print(f"  NOT IN ENSEMBL — needs manual lookup")
            for sp in TARGET_SPECIES:
                name = sp["name"]
                type_col = col_map[name][f"{name}_orthologue_type"]
                ws.cell(row=row, column=type_col, value="NOT_IN_ENSEMBL")
                summary[name]["not_in_ensembl"] += 1
            continue

        print(f"  Ensembl ID: {ensembl_id}")

        for sp in TARGET_SPECIES:
            name = sp["name"]
            ensembl_sp = sp["ensembl"]

            time.sleep(0.4)

            orthos = get_orthologues(ensembl_id, ensembl_sp)

            if not orthos:
                print(f"  {name}: ABSENT")
                ws.cell(row=row, column=col_map[name][f"{name}_orthologue_type"],
                        value="ABSENT")
                summary[name]["absent"] += 1
            else:
                best = orthos[0]
                print(f"  {name}: {best['type']} → {best['target_symbol']} "
                      f"({best['pct_id_query']}% id, {best['pct_pos']}% pos)")

                ws.cell(row=row, column=col_map[name][f"{name}_orthologue_type"],
                        value=best["type"])
                ws.cell(row=row, column=col_map[name][f"{name}_target_id"],
                        value=best["target_id"])
                ws.cell(row=row, column=col_map[name][f"{name}_target_symbol"],
                        value=best["target_symbol"])
                ws.cell(row=row, column=col_map[name][f"{name}_pct_id"],
                        value=best["pct_id_query"])
                ws.cell(row=row, column=col_map[name][f"{name}_pct_pos"],
                        value=best["pct_pos"])
                ws.cell(row=row, column=col_map[name][f"{name}_confidence"],
                        value=str(best["is_high_confidence"]))
                ws.cell(row=row, column=col_map[name][f"{name}_dn_ds"],
                        value=str(best["dn_ds"]) if best["dn_ds"] else "")

                if len(orthos) > 1:
                    current = ws.cell(row=row, column=col_map[name][f"{name}_orthologue_type"]).value
                    ws.cell(row=row, column=col_map[name][f"{name}_orthologue_type"],
                            value=f"{current} (+{len(orthos)-1} more)")

                summary[name]["present"] += 1

        time.sleep(0.3)

    wb.save(output_path)

    print(f"\n{'='*60}")
    print(f"Saved: {output_path}")
    print(f"Total genes: {total}")
    for sp in TARGET_SPECIES:
        name = sp["name"]
        s = summary[name]
        print(f"\n{name.upper()}:")
        print(f"  Present:        {s['present']}")
        print(f"  Absent:         {s['absent']}")
        print(f"  Not in Ensembl: {s['not_in_ensembl']}")


if __name__ == "__main__":
    main()
