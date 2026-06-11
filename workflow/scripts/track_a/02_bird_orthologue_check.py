#!/usr/bin/env python3
"""
02_bird_orthologue_check.py
===========================
For each human gene in the table, query Ensembl Compara for orthologues in:
  - Chicken (Gallus gallus)
  - Duck (Anas platyrhynchos)

This is a FAST FIRST PASS using pre-computed orthologues. It is NOT a BLAST
search — it queries Ensembl's orthologue database, which uses gene trees
built from whole-genome alignments. This is Step 0.5 in the pipeline:
filling the "Bird Status" column before you run the proper BLAST (Step 2).

For each species, reports:
  - Orthologue type (one2one, one2many, many2many, or absent)
  - Target gene ID and symbol
  - Percent identity (query→target and target→query)
  - Confidence (high/low from Ensembl's gene order conservation score)

Outputs:
  - Updated xlsx with new columns per species
  - Console summary of presence/absence

Requirements:
  pip install requests openpyxl

Usage:
  python 02_bird_orthologue_check.py human_cell_death_gene_universe_v1.xlsx

Notes:
  - Ensembl REST API is free, no authentication needed
  - Rate limit: 15 requests/second (we use 0.3s delay to be safe)
  - Uses Ensembl gene symbols to map from our table
  - Some genes may not be in Ensembl (e.g. if symbol differs); these are
    flagged for manual follow-up
"""

import sys
import time
import json
import requests
import openpyxl
from pathlib import Path


ENSEMBL_REST = "https://rest.ensembl.org"

# Species we care about
TARGET_SPECIES = [
    {"name": "chicken", "ensembl": "gallus_gallus"},
    {"name": "duck", "ensembl": "anas_platyrhynchos"},
]


def ensembl_get(endpoint: str, max_retries: int = 3) -> dict | None:
    """GET request to Ensembl REST API with retry logic."""
    url = f"{ENSEMBL_REST}{endpoint}"
    headers = {"Content-Type": "application/json"}
    for attempt in range(max_retries):
        try:
            r = requests.get(url, headers=headers, timeout=20)
            if r.status_code == 200:
                return r.json()
            elif r.status_code == 429:
                wait = float(r.headers.get("Retry-After", 2))
                print(f"    Rate limited, waiting {wait}s...")
                time.sleep(wait)
            elif r.status_code == 400:
                return None  # bad request, gene not found
            else:
                time.sleep(1)
        except requests.exceptions.RequestException as e:
            print(f"    Request error: {e}")
            time.sleep(2)
    return None


def lookup_human_gene(symbol: str) -> str | None:
    """Get the Ensembl gene ID for a human gene symbol."""
    data = ensembl_get(f"/lookup/symbol/homo_sapiens/{symbol}?expand=0")
    if data and "id" in data:
        return data["id"]
    return None


def get_orthologues(ensembl_id: str, target_species: str) -> list[dict]:
    """Get orthologues for a human Ensembl gene ID in a target species."""
    data = ensembl_get(
        f"/homology/id/{ensembl_id}?"
        f"target_species={target_species}&type=orthologues"
    )
    if not data:
        return []

    homologies = data.get("data", [{}])[0].get("homologies", [])
    results = []
    for h in homologies:
        target = h.get("target", {})
        results.append({
            "type": h.get("type", "unknown"),
            "target_id": target.get("id", ""),
            "target_symbol": target.get("gene_member_id", ""),
            "target_protein": target.get("protein_id", ""),
            "pct_id_query": h.get("source", {}).get("perc_id", ""),
            "pct_id_target": target.get("perc_id", ""),
            "confidence": h.get("taxonomy_level", ""),
            "dn_ds": h.get("dn_ds", ""),
            "is_high_confidence": h.get("is_high_confidence", ""),
        })
    return results


def try_aliases(symbol: str, aliases: str) -> str | None:
    """Try the primary symbol first, then any aliases."""
    # Try primary symbol
    eid = lookup_human_gene(symbol)
    if eid:
        return eid

    # Try aliases
    if aliases:
        for alias in aliases.split(","):
            alias = alias.strip()
            if alias:
                eid = lookup_human_gene(alias)
                if eid:
                    print(f"    (found via alias: {alias})")
                    return eid
    return None


def main():
    if len(sys.argv) < 2:
        print("Usage: python 02_bird_orthologue_check.py <input.xlsx>")
        sys.exit(1)

    input_path = Path(sys.argv[1])
    output_path = input_path.with_stem(input_path.stem + "_bird_orthologues")

    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    # Find column indices
    headers = {cell.value: cell.column for cell in ws[1]}
    col_symbol = headers.get("Gene Symbol")
    col_aliases = headers.get("Aliases")

    if not col_symbol:
        print("ERROR: Could not find 'Gene Symbol' column")
        sys.exit(1)

    # Add new columns for each target species
    start_col = ws.max_column + 1
    col_map = {}
    for sp in TARGET_SPECIES:
        name = sp["name"]
        cols_for_sp = [
            f"{name}_orthologue_type",
            f"{name}_target_id",
            f"{name}_target_symbol",
            f"{name}_pct_id",
            f"{name}_confidence",
            f"{name}_dn_ds",
        ]
        col_map[name] = {}
        for i, cname in enumerate(cols_for_sp):
            col_idx = start_col
            ws.cell(row=1, column=col_idx, value=cname)
            ws.cell(row=1, column=col_idx).font = openpyxl.styles.Font(bold=True)
            col_map[name][cname] = col_idx
            start_col += 1

    # Process each gene
    total = ws.max_row - 1
    summary = {sp["name"]: {"present": 0, "absent": 0, "not_in_ensembl": 0} for sp in TARGET_SPECIES}

    for row in range(2, ws.max_row + 1):
        symbol = (ws.cell(row=row, column=col_symbol).value or "").strip()
        aliases = (ws.cell(row=row, column=col_aliases).value or "").strip() if col_aliases else ""

        print(f"\n[{row-1}/{total}] {symbol}")

        # Look up Ensembl ID
        ensembl_id = try_aliases(symbol, aliases)

        if not ensembl_id:
            print(f"  NOT IN ENSEMBL — needs manual lookup")
            for sp in TARGET_SPECIES:
                name = sp["name"]
                type_col = col_map[name][f"{name}_orthologue_type"]
                ws.cell(row=row, column=type_col, value="NOT_IN_ENSEMBL")
                summary[name]["not_in_ensembl"] += 1
            continue

        print(f"  Ensembl ID: {ensembl_id}")

        # Check each target species
        for sp in TARGET_SPECIES:
            name = sp["name"]
            ensembl_sp = sp["ensembl"]

            time.sleep(0.4)  # rate limiting

            orthos = get_orthologues(ensembl_id, ensembl_sp)

            if not orthos:
                print(f"  {name}: ABSENT")
                type_col = col_map[name][f"{name}_orthologue_type"]
                ws.cell(row=row, column=type_col, value="ABSENT")
                summary[name]["absent"] += 1
            else:
                # Take the first (highest-confidence) orthologue
                best = orthos[0]
                print(f"  {name}: {best['type']} → {best['target_symbol']} "
                      f"({best['pct_id_query']}% identity)")

                ws.cell(row=row, column=col_map[name][f"{name}_orthologue_type"],
                        value=best["type"])
                ws.cell(row=row, column=col_map[name][f"{name}_target_id"],
                        value=best["target_id"])
                ws.cell(row=row, column=col_map[name][f"{name}_target_symbol"],
                        value=best["target_symbol"])
                ws.cell(row=row, column=col_map[name][f"{name}_pct_id"],
                        value=best["pct_id_query"])
                ws.cell(row=row, column=col_map[name][f"{name}_confidence"],
                        value=str(best["is_high_confidence"]))
                ws.cell(row=row, column=col_map[name][f"{name}_dn_ds"],
                        value=str(best["dn_ds"]) if best["dn_ds"] else "")

                # If multiple orthologues, note it
                if len(orthos) > 1:
                    current = ws.cell(row=row, column=col_map[name][f"{name}_orthologue_type"]).value
                    ws.cell(row=row, column=col_map[name][f"{name}_orthologue_type"],
                            value=f"{current} (+{len(orthos)-1} more)")

                summary[name]["present"] += 1

        time.sleep(0.3)

    # Save
    wb.save(output_path)

    # Print summary
    print(f"\n{'='*60}")
    print(f"Saved: {output_path}")
    print(f"Total genes: {total}")
    for sp in TARGET_SPECIES:
        name = sp["name"]
        s = summary[name]
        print(f"\n{name.upper()}:")
        print(f"  Present:        {s['present']}")
        print(f"  Absent:         {s['absent']}")
        print(f"  Not in Ensembl: {s['not_in_ensembl']}")


if __name__ == "__main__":
    main()
