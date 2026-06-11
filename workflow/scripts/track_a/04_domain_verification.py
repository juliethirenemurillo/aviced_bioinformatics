#!/usr/bin/env python3
"""
04_domain_verification.py
=========================
For each bird orthologue found by Ensembl Compara, verify that it has
the same functional domains as the human reference protein.

Workflow per gene:
  1. Get human protein's InterPro domains (from UniProt accession)
  2. Map bird Ensembl gene ID → UniProt accession (via Ensembl xrefs)
  3. Get bird protein's InterPro domains
  4. Compare: which human domains are present/absent in the bird protein?

Output: updated xlsx with domain verification columns per species.

Requirements:
  pip install requests openpyxl

Usage:
  python 04_domain_verification.py human_cell_death_gene_universe_v1_bird_orthologues.xlsx
"""

import sys
import time
import json
import requests
import openpyxl
from pathlib import Path
from collections import defaultdict


ENSEMBL_REST = "https://rest.ensembl.org"
UNIPROT_REST = "https://rest.uniprot.org"


# ── API helpers ──────────────────────────────────────────────────────

def api_get(url: str, headers: dict = None, params: dict = None,
            max_retries: int = 3, delay: float = 0.5):
    """Generic GET with retry logic."""
    if headers is None:
        headers = {"Accept": "application/json"}
    for attempt in range(max_retries):
        try:
            r = requests.get(url, headers=headers, params=params, timeout=25)
            if r.status_code == 200:
                return r.json()
            elif r.status_code == 429:
                wait = float(r.headers.get("Retry-After", 3))
                print(f"      Rate limited, waiting {wait}s...")
                time.sleep(wait)
            elif r.status_code == 404:
                return None
            else:
                if attempt < max_retries - 1:
                    time.sleep(1)
        except requests.exceptions.RequestException as e:
            if attempt < max_retries - 1:
                time.sleep(2)
            else:
                print(f"      Request failed: {e}")
    return None


# ── Domain extraction ────────────────────────────────────────────────

def get_interpro_domains_from_uniprot(uniprot_accession: str) -> list[dict]:
    """
    Get InterPro domain annotations for a UniProt accession.
    Returns list of dicts: {id, name, type} where type is e.g. 'Domain', 'Family', 'Repeat'
    """
    data = api_get(f"{UNIPROT_REST}/uniprotkb/{uniprot_accession}.json")
    if not data:
        return []

    domains = []
    seen = set()
    xrefs = data.get("uniProtKBCrossReferences", [])
    for xref in xrefs:
        if xref.get("database") == "InterPro":
            ipr_id = xref.get("id", "")
            if ipr_id in seen:
                continue
            seen.add(ipr_id)
            props = {p["key"]: p["value"] for p in xref.get("properties", [])}
            domains.append({
                "id": ipr_id,
                "name": props.get("EntryName", ""),
                "type": props.get("Type", ""),  # not always present here
            })

    return domains


def get_uniprot_from_ensembl(ensembl_gene_id: str, species: str) -> str | None:
    """
    Map an Ensembl gene ID to a UniProt accession via Ensembl xrefs.
    Prefers Swiss-Prot; falls back to TrEMBL.
    """
    data = api_get(
        f"{ENSEMBL_REST}/xrefs/id/{ensembl_gene_id}",
        params={"external_db": "Uniprot_gn", "all_levels": "1"}
    )
    if not data:
        # Try broader search without external_db filter
        data = api_get(f"{ENSEMBL_REST}/xrefs/id/{ensembl_gene_id}")
        if not data:
            return None

    # Collect UniProt hits
    swissprot = []
    trembl = []
    for entry in data:
        dbname = entry.get("dbname", "")
        accession = entry.get("primary_id", "")
        if not accession:
            continue
        if "Uniprot/SWISSPROT" in dbname or "UniProtKB/Swiss-Prot" in dbname:
            swissprot.append(accession)
        elif "Uniprot/SPTREMBL" in dbname or "UniProtKB/TrEMBL" in dbname:
            trembl.append(accession)
        elif "Uniprot" in dbname:
            trembl.append(accession)

    # Prefer Swiss-Prot
    if swissprot:
        return swissprot[0]
    if trembl:
        return trembl[0]
    return None


# ── Domain comparison ────────────────────────────────────────────────

def compare_domains(human_domains: list[dict], bird_domains: list[dict]) -> dict:
    """
    Compare human vs bird InterPro domain sets.
    Returns: {
        shared: [ids present in both],
        human_only: [ids in human but not bird],
        bird_only: [ids in bird but not human],
        human_count: int,
        bird_count: int,
        match_pct: float (shared / human * 100),
        verdict: str
    }
    """
    human_ids = {d["id"] for d in human_domains}
    bird_ids = {d["id"] for d in bird_domains}

    shared = human_ids & bird_ids
    human_only = human_ids - bird_ids
    bird_only = bird_ids - human_ids

    # Human domain names for reporting
    human_id_to_name = {d["id"]: d["name"] for d in human_domains}
    bird_id_to_name = {d["id"]: d["name"] for d in bird_domains}

    match_pct = (len(shared) / len(human_ids) * 100) if human_ids else 0

    if match_pct == 100:
        verdict = "FULL_MATCH"
    elif match_pct >= 75:
        verdict = "GOOD_MATCH"
    elif match_pct >= 50:
        verdict = "PARTIAL_MATCH"
    elif match_pct > 0:
        verdict = "WEAK_MATCH"
    else:
        verdict = "NO_MATCH"

    return {
        "shared": [f"{ipr}:{human_id_to_name.get(ipr, '')}" for ipr in sorted(shared)],
        "human_only": [f"{ipr}:{human_id_to_name.get(ipr, '')}" for ipr in sorted(human_only)],
        "bird_only": [f"{ipr}:{bird_id_to_name.get(ipr, '')}" for ipr in sorted(bird_only)],
        "human_count": len(human_ids),
        "bird_count": len(bird_ids),
        "match_pct": round(match_pct, 1),
        "verdict": verdict,
    }


# ── Main ─────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print("Usage: python 04_domain_verification.py <bird_orthologues.xlsx>")
        sys.exit(1)

    input_path = Path(sys.argv[1])
    output_path = input_path.with_stem(input_path.stem + "_domains")

    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    headers = {cell.value: cell.column for cell in ws[1]}
    col_symbol = headers.get("Gene Symbol")
    col_human_uniprot = headers.get("UniProt Accession")
    col_chk_id = headers.get("chicken_target_id")
    col_chk_type = headers.get("chicken_orthologue_type")
    col_duck_id = headers.get("duck_target_id")
    col_duck_type = headers.get("duck_orthologue_type")

    if not col_symbol or not col_human_uniprot:
        print("ERROR: Missing required columns")
        print(f"Available: {list(headers.keys())}")
        sys.exit(1)

    # Add new columns
    species_cols = []
    for sp in ["chicken", "duck"]:
        sp_cols = [
            f"{sp}_uniprot",
            f"{sp}_domain_verdict",
            f"{sp}_domain_match_pct",
            f"{sp}_shared_domains",
            f"{sp}_missing_domains",
            f"{sp}_extra_domains",
        ]
        species_cols.append((sp, sp_cols))

    start_col = ws.max_column + 1
    col_map = {}
    for sp, cols in species_cols:
        col_map[sp] = {}
        for cname in cols:
            ws.cell(row=1, column=start_col, value=cname)
            ws.cell(row=1, column=start_col).font = openpyxl.styles.Font(bold=True)
            col_map[sp][cname] = start_col
            start_col += 1

    total = ws.max_row - 1

    # Cache human domains to avoid re-querying
    human_domain_cache = {}

    for row in range(2, ws.max_row + 1):
        symbol = (ws.cell(row=row, column=col_symbol).value or "").strip()
        human_acc = (ws.cell(row=row, column=col_human_uniprot).value or "").strip()

        print(f"\n[{row-1}/{total}] {symbol}")

        # Get human domains (cached)
        if human_acc not in human_domain_cache:
            print(f"  Human ({human_acc}): fetching domains...", end=" ")
            time.sleep(0.3)
            human_domains = get_interpro_domains_from_uniprot(human_acc)
            human_domain_cache[human_acc] = human_domains
            print(f"{len(human_domains)} InterPro entries")
        else:
            human_domains = human_domain_cache[human_acc]

        # Process each species
        for sp, sp_col_names in species_cols:
            sp_name = sp
            id_col = col_chk_id if sp == "chicken" else col_duck_id
            type_col = col_chk_type if sp == "chicken" else col_duck_type

            if not id_col:
                continue

            bird_ensembl_id = (ws.cell(row=row, column=id_col).value or "").strip()
            bird_type = (ws.cell(row=row, column=type_col).value or "").strip()

            # Skip absent genes
            if not bird_ensembl_id or bird_type in ("ABSENT", "NOT_IN_ENSEMBL", "—", ""):
                ws.cell(row=row, column=col_map[sp][f"{sp}_domain_verdict"],
                        value="NO_ORTHOLOGUE")
                continue

            print(f"  {sp_name} ({bird_ensembl_id}):", end=" ")

            # Map Ensembl → UniProt
            time.sleep(0.4)
            bird_uniprot = get_uniprot_from_ensembl(bird_ensembl_id, sp_name)

            if not bird_uniprot:
                print("no UniProt mapping")
                ws.cell(row=row, column=col_map[sp][f"{sp}_uniprot"],
                        value="NOT_FOUND")
                ws.cell(row=row, column=col_map[sp][f"{sp}_domain_verdict"],
                        value="NO_UNIPROT")
                continue

            print(f"UniProt={bird_uniprot}", end=" → ")
            ws.cell(row=row, column=col_map[sp][f"{sp}_uniprot"],
                    value=bird_uniprot)

            # Get bird domains
            time.sleep(0.3)
            bird_domains = get_interpro_domains_from_uniprot(bird_uniprot)
            print(f"{len(bird_domains)} domains", end=" → ")

            # Compare
            comparison = compare_domains(human_domains, bird_domains)
            print(f"{comparison['verdict']} ({comparison['match_pct']}%)")

            ws.cell(row=row, column=col_map[sp][f"{sp}_domain_verdict"],
                    value=comparison["verdict"])
            ws.cell(row=row, column=col_map[sp][f"{sp}_domain_match_pct"],
                    value=comparison["match_pct"])
            ws.cell(row=row, column=col_map[sp][f"{sp}_shared_domains"],
                    value="; ".join(comparison["shared"]) if comparison["shared"] else "—")
            ws.cell(row=row, column=col_map[sp][f"{sp}_missing_domains"],
                    value="; ".join(comparison["human_only"]) if comparison["human_only"] else "—")
            ws.cell(row=row, column=col_map[sp][f"{sp}_extra_domains"],
                    value="; ".join(comparison["bird_only"]) if comparison["bird_only"] else "—")

    wb.save(output_path)

    print(f"\n{'='*60}")
    print(f"Saved: {output_path}")
    print(f"\nVERDICT KEY:")
    print(f"  FULL_MATCH    = all human InterPro domains found in bird protein")
    print(f"  GOOD_MATCH    = ≥75% of human domains present")
    print(f"  PARTIAL_MATCH = 50-75% of human domains present")
    print(f"  WEAK_MATCH    = <50% of human domains present")
    print(f"  NO_MATCH      = no shared domains")
    print(f"  NO_ORTHOLOGUE = gene absent in this species")
    print(f"  NO_UNIPROT    = Ensembl ID couldn't be mapped to UniProt")


if __name__ == "__main__":
    main()
